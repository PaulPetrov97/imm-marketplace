"""Generator pentru workbook-ul 'Analiza incadrare IMM' (format RBC contractare).

LIVRABIL OBLIGATORIU la fiecare analiza. Produce un raport complet intr-un
singur sheet, cu:
  1. Harta grupului  — per firma: denumire+CUI, asociati, cota, administrator,
     relatii comerciale (DA/NU), adresa, CAEN principal, CAEN preponderent.
     Sursa: termene.ro — AMBELE sectiuni: "Asociati/actionari" SI
     "Persoane autorizate" (administratorii pot lega firme fara participatie).
  2. Verdict legaturi — LEGATA/PARTENERA/AUTONOMA/NECLAR + lista firmelor
  3. Tabele financiare per an — angajati, CA si Active totale in LEI, EURO,
     MII LEI si MII EURO (coloanele derivate raman formule live)
  4. Categoria finala IMM (micro/mica/mijlocie/nu_imm)
  5. Recomandare Claude — concluzia consultantului AI: ce s-a constatat,
     ce ramane de clarificat, ce varianta de declarare se recomanda.

REGULI DE FORMAT (cerinte RBC):
  * Cursul EUR/RON = cursul BNR din ULTIMA ZI (lucratoare) a anului de
    referinta (31 decembrie). Vezi reference/06-format-analiza-imm.md.
  * Separatori romanesti: zecimale cu virgula, mii cu punct. In Excel,
    codurile de format (#,##0.00) sunt redate conform setarilor regionale —
    pe Windows cu locale Romania se afiseaza automat 1.234.567,89.
  * Coloanele euro = formule live  =lei/curs ; mii = formule  /1000.
  * TOTAL sumeaza DOAR intreprinderile luate in calcul (randurile [EXCLUS]
    apar in tabel dar nu intra in TOTAL).

API:
    build_analiza_imm(out_path, solicitant, companies, verdict, years,
                      categorie, recomandare="", program=None)
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# -------------------- Stiluri --------------------
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_SUBHDR_FILL = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
_VERDICT_FILL = PatternFill("solid", fgColor="C6E0B4")
_RECOM_FILL = PatternFill("solid", fgColor="FFF2CC")
_BOLD = Font(bold=True, size=10)
_NORMAL = Font(size=10)
_thin = Side(style="thin", color="999999")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Format numeric: separatorii sunt redati conform locale-ului Windows/Excel
# (pe sisteme RO: mii cu '.', zecimale cu ',').
_FMT_LEI = "#,##0.00"
_FMT_MII = "#,##0.000"


def fmt_ro(value: float, decimals: int = 2) -> str:
    """Numar in conventia romaneasca: 1.234.567,89 (pentru texte, nu Excel)."""
    s = f"{value:,.{decimals}f}"
    return s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _style_row(ws, row, c0, c1, *, fill=None, font=None, align=None, border=True):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if border:
            cell.border = _BORDER


def build_analiza_imm(
    out_path: Path,
    solicitant: dict,
    companies: list[dict],
    verdict: dict,
    years: list[dict],
    categorie: str,
    recomandare: str = "",
    program: str | None = None,
) -> Path:
    """
    solicitant: {denumire, cui}
    companies:  list de blocuri, fiecare:
        {denumire, cui, asociati: [{nume, cota}], administrator,
         relatii: "DA"/"NU", adresa, caen_principal, caen_preponderent}
        -> include si administratorii/persoanele autorizate relevante!
    verdict:    {tip: "LEGATA"/"PARTENERA"/"AUTONOMA"/"NECLAR", cu: "text firme",
                 nota: "text optional"}
    years:      list de {an, curs, rows: [{denumire, angajati, ca_lei, active_lei,
                 exclus: bool, nota: str}]}
                curs = cursul BNR din ULTIMA ZI a anului respectiv (31 dec)!
    categorie:  "MICROINTREPRINDERE"/"MICA"/"MIJLOCIE"/"NU ESTE IMM"
    recomandare: textul "Recomandare Claude" (concluzie + pasi urmatori)
    program:    programul / apelul de finantare (optional, apare in antet)
    """
    wb = Workbook()
    ws = wb.active
    title = solicitant["denumire"][:31]
    ws.title = "".join(ch for ch in title if ch not in r'\/?*[]:')[:31] or "Analiza IMM"

    # Latimi coloane (B..G harta grupului; B..K tabele financiare; M..N curs/note)
    widths = {"A": 2, "B": 34, "C": 28, "D": 16, "E": 24, "F": 17, "G": 36,
              "H": 14, "I": 16, "J": 14, "K": 16, "L": 2, "M": 28, "N": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    NET_HDRS = ["Denumire societate", "Asociati", "Cota de participare",
                "Administrator", "Relatii comerciale / de alta natura", "Adresa"]

    r = 2
    ws.cell(row=r, column=2, value="ANALIZA INCADRARE IMM").font = Font(bold=True, size=13, color="1F4E79")
    r += 1
    ws.cell(row=r, column=2,
            value=f"Solicitant: {solicitant['denumire']} (CUI {solicitant['cui']})").font = _BOLD
    r += 1
    if program:
        ws.cell(row=r, column=2, value=f"Program / apel de finantare: {program}").font = _NORMAL
        r += 1
    ws.cell(row=r, column=2,
            value=f"Data analizei: {date.today().strftime('%d.%m.%Y')} · "
                  f"Surse: termene.ro (Asociati/actionari + Persoane autorizate) + ONRC/ANAF"
            ).font = Font(italic=True, size=9)
    r += 2

    # -------------------- 1. Harta grupului --------------------
    ws.cell(row=r, column=2, value="1. STRUCTURA GRUPULUI").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    for comp in companies:
        for i, h in enumerate(NET_HDRS):
            ws.cell(row=r, column=2 + i, value=h)
        _style_row(ws, r, 2, 7, fill=_SUBHDR_FILL, font=_BOLD, align=_CENTER)
        r += 1
        asociati = comp.get("asociati", [{}]) or [{}]
        first = r
        for j, a in enumerate(asociati):
            ws.cell(row=r, column=2, value=comp["denumire"] + (f" {comp['cui']}" if comp.get("cui") and j == 0 else "") if j == 0 else "")
            ws.cell(row=r, column=3, value=a.get("nume", ""))
            cota = a.get("cota", "")
            cell_d = ws.cell(row=r, column=4, value=cota)
            if isinstance(cota, (int, float)):
                cell_d.number_format = "0.00%"
            ws.cell(row=r, column=5, value=comp.get("administrator", "") if j == 0 else "")
            ws.cell(row=r, column=6, value=comp.get("relatii", "") if j == 0 else "")
            ws.cell(row=r, column=7, value=comp.get("adresa", "") if j == 0 else "")
            _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
            r += 1
        if len(asociati) > 1:
            for col in (2, 5, 6, 7):
                ws.merge_cells(start_row=first, start_column=col, end_row=r - 1, end_column=col)
        ws.cell(row=r, column=2, value="CAEN principal")
        ws.cell(row=r, column=3, value=comp.get("caen_principal", ""))
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
        r += 1
        ws.cell(row=r, column=2, value="CAEN preponderent")
        ws.cell(row=r, column=3, value=comp.get("caen_preponderent", ""))
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
        r += 2

    # -------------------- 2. Verdict --------------------
    ws.cell(row=r, column=2, value="2. CONCLUZIE LEGATURI").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    ws.cell(row=r, column=2, value=verdict.get("tip", ""))
    ws.cell(row=r, column=3, value=verdict.get("cu", ""))
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
    _style_row(ws, r, 2, 11, fill=_VERDICT_FILL, font=_BOLD, align=_WRAP)
    r += 1
    if verdict.get("nota"):
        ws.cell(row=r, column=2, value="Nota")
        ws.cell(row=r, column=3, value=verdict["nota"])
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        _style_row(ws, r, 2, 11, font=Font(italic=True, size=9), align=_WRAP)
        ws.row_dimensions[r].height = max(28, 13 * (len(verdict["nota"]) // 130 + 1))
        r += 1
    r += 1

    # -------------------- 3. Tabele financiare per an --------------------
    ws.cell(row=r, column=2, value="3. DATE FINANCIARE CONSOLIDATE (per an)").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    FIN_HDRS = ["Denumire societate", "Nr. angajati", "CA (lei)",
                "Active totale (lei)", "CA (euro)", "Active totale (euro)",
                "CA (mii lei)", "Active totale (mii lei)",
                "CA (mii euro)", "Active totale (mii euro)"]
    for yd in years:
        an = yd["an"]
        curs = yd["curs"]
        ws.cell(row=r, column=2, value=f"Exercitiul financiar {an}").font = _BOLD
        ws.cell(row=r, column=13, value=f"curs euro BNR 31.12.{an}").font = _NORMAL
        jcell = ws.cell(row=r, column=14, value=curs)
        jcell.number_format = "0.0000"
        r += 1
        for i, h in enumerate(FIN_HDRS):
            ws.cell(row=r, column=2 + i, value=h)
        _style_row(ws, r, 2, 11, fill=_SUBHDR_FILL, font=_BOLD, align=_CENTER)
        r += 1
        sums = {c: 0.0 for c in range(3, 12)}  # C..K — totaluri numerice
        for row in yd["rows"]:
            excl = bool(row.get("exclus"))
            ca_lei = float(row.get("ca_lei", 0) or 0)
            act_lei = float(row.get("active_lei", 0) or 0)
            ca_eur = ca_lei / curs if curs else 0.0
            act_eur = act_lei / curs if curs else 0.0
            ang = row.get("angajati", "-")
            # euro & mii = NUMERE calculate (nu formule) — vizibile in orice viewer/PDF
            vals = {
                4: round(ca_lei, 2), 5: round(act_lei, 2),
                6: round(ca_eur, 2), 7: round(act_eur, 2),
                8: round(ca_lei / 1000, 3), 9: round(act_lei / 1000, 3),
                10: round(ca_eur / 1000, 3), 11: round(act_eur / 1000, 3),
            }
            ws.cell(row=r, column=2, value=row["denumire"] + (" [EXCLUS]" if excl else ""))
            ws.cell(row=r, column=3, value=ang)
            for idx, v in vals.items():
                cell = ws.cell(row=r, column=idx, value=v)
                cell.number_format = _FMT_MII if idx >= 8 else _FMT_LEI
            if not excl:
                if isinstance(ang, (int, float)):
                    sums[3] += ang
                for idx, v in vals.items():
                    sums[idx] += v
            _style_row(ws, r, 2, 11, font=_NORMAL, align=_WRAP)
            if row.get("nota"):
                ws.cell(row=r, column=13, value=row["nota"]).font = Font(italic=True, size=8)
            r += 1
        # TOTAL = suma NUMERICA a randurilor luate in calcul (exclude [EXCLUS])
        ws.cell(row=r, column=2, value="TOTAL (intreprinderi luate in calcul)")
        ws.cell(row=r, column=3, value=int(round(sums[3])))
        for idx in range(4, 12):
            cc = ws.cell(row=r, column=idx, value=round(sums[idx], 3 if idx >= 8 else 2))
            cc.number_format = _FMT_MII if idx >= 8 else _FMT_LEI
        _style_row(ws, r, 2, 11, fill=_TOTAL_FILL, font=_BOLD, align=_CENTER)
        r += 2

    # -------------------- 4. Categoria finala --------------------
    ws.cell(row=r, column=2, value="4. CATEGORIA IMM").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    ws.cell(row=r, column=2, value="Incadrare finala:")
    cell_cat = ws.cell(row=r, column=3, value=categorie)
    cell_cat.font = Font(bold=True, size=12, color="C00000")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    _style_row(ws, r, 2, 6, font=_BOLD, align=_CENTER)
    r += 2

    # -------------------- 5. Recomandare Claude --------------------
    ws.cell(row=r, column=2, value="5. RECOMANDARE CLAUDE").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    rec = recomandare or "(fara recomandare)"
    ws.cell(row=r, column=2, value=rec)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    _style_row(ws, r, 2, 11, fill=_RECOM_FILL, font=_NORMAL, align=_WRAP)
    ws.row_dimensions[r].height = max(45, 13 * (rec.count("\n") + len(rec) // 150 + 1))
    r += 1
    ws.cell(row=r, column=2,
            value=f"Generat automat de pluginul imm-analiza-ro · {date.today().strftime('%d.%m.%Y')} · "
                  f"A se valida cu declaratiile pe propria raspundere ale solicitantului."
            ).font = Font(italic=True, size=8, color="808080")

    ws.sheet_view.showGridLines = False
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    # Smoke minimal
    out = Path.cwd() / "_smoke_analiza_imm.xlsx"
    build_analiza_imm(
        out,
        {"denumire": "DEMO SRL", "cui": "12345678"},
        [{"denumire": "DEMO SRL", "cui": "12345678",
          "asociati": [{"nume": "POP ION", "cota": 1.0}],
          "administrator": "POP ION", "relatii": "-",
          "adresa": "Bucuresti", "caen_principal": "6201", "caen_preponderent": "6201"}],
        {"tip": "AUTONOMA", "cu": "-", "nota": "demo"},
        [{"an": 2024, "curs": 4.9741,
          "rows": [{"denumire": "DEMO SRL", "angajati": 5, "ca_lei": 1234567.89, "active_lei": 500000}]}],
        "MICROINTREPRINDERE",
        recomandare="Demo: firma autonoma, micro. Se recomanda declararea ca autonoma; "
                    "nu exista cazuri NECLAR.",
        program="Demo apel 2026",
    )
    print(f"OK: {out}")
    print(f"fmt_ro test: {fmt_ro(1234567.891, 2)} (asteptat 1.234.567,89)")
