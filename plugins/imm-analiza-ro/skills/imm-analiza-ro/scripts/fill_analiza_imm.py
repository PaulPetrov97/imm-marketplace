"""Generator pentru workbook-ul 'Analiza incadrare IMM' (format Hexol Lube).

Produce un raport complet de analiza IMM intr-un singur sheet, cu:
  1. Harta grupului  — per firma: denumire+CUI, asociati, cota, administrator,
     relatii comerciale (DA/NU), adresa, CAEN principal, CAEN preponderent
  2. Verdict legaturi — LEGATA/PARTENERA/AUTONOMA + lista firmelor
  3. Tabele financiare per an — angajati, CA (lei), Active totale (lei),
     CA (euro)=lei/curs, Active totale (euro)=lei/curs, TOTAL pe coloane
  4. Categoria finala IMM (micro/mica/mijlocie/nu_imm)

Spre deosebire de Declaratie_IMM_v1.0.xlsx (template fix cu max 8 parteneri),
acesta se construieste de la zero (openpyxl) si scaleaza la orice numar de firme
si ani. Coloanele euro raman formule live (=lei/curs).

API:
    build_analiza_imm(out_path, solicitant, companies, verdict, years, categorie)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# -------------------- Stiluri --------------------
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_SUBHDR_FILL = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
_VERDICT_FILL = PatternFill("solid", fgColor="C6E0B4")
_BOLD = Font(bold=True, size=10)
_NORMAL = Font(size=10)
_thin = Side(style="thin", color="999999")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _style_row(ws, row, c0, c1, *, fill=None, font=None, align=None, border=True):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if border:
            cell.border = _BORDER


def build_analiza_imm(
    out_path: Path,
    solicitant: dict,
    companies: list[dict],
    verdict: dict,
    years: list[dict],
    categorie: str,
) -> Path:
    """
    solicitant: {denumire, cui}
    companies:  list de blocuri, fiecare:
        {denumire, cui, asociati: [{nume, cota}], administrator,
         relatii: "DA"/"NU", adresa, caen_principal, caen_preponderent}
    verdict:    {tip: "LEGATA"/"PARTENERA"/"AUTONOMA"/"NECLAR", cu: "text firme",
                 nota: "text optional"}
    years:      list de {an, curs, rows: [{denumire, angajati, ca_lei, active_lei,
                 exclus: bool, nota: str}]}
    categorie:  "MICROINTREPRINDERE"/"MICA"/"MIJLOCIE"/"NU ESTE IMM"
    """
    wb = Workbook()
    ws = wb.active
    title = solicitant["denumire"][:31]
    ws.title = "".join(ch for ch in title if ch not in r'\/?*[]:')[:31] or "Analiza IMM"

    # Latimi coloane
    widths = {"A": 2, "B": 34, "C": 30, "D": 16, "E": 26, "F": 18, "G": 40,
              "H": 2, "I": 12, "J": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    NET_HDRS = ["Denumire societate", "Asociati", "Cota de participare",
                "Administrator", "Relatii comerciale / de alta natura", "Adresa"]

    r = 2
    ws.cell(row=r, column=2, value="ANALIZA INCADRARE IMM").font = Font(bold=True, size=13, color="1F4E79")
    r += 1
    ws.cell(row=r, column=2,
            value=f"Solicitant: {solicitant['denumire']} (CUI {solicitant['cui']})").font = _BOLD
    r += 2

    # -------------------- 1. Harta grupului --------------------
    ws.cell(row=r, column=2, value="1. STRUCTURA GRUPULUI").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    for comp in companies:
        # Header
        for i, h in enumerate(NET_HDRS):
            ws.cell(row=r, column=2 + i, value=h)
        _style_row(ws, r, 2, 7, fill=_SUBHDR_FILL, font=_BOLD, align=_CENTER)
        r += 1
        asociati = comp.get("asociati", [{}]) or [{}]
        first = r
        for j, a in enumerate(asociati):
            ws.cell(row=r, column=2, value=comp["denumire"] + (f" {comp['cui']}" if comp.get("cui") and j == 0 else "") if j == 0 else "")
            ws.cell(row=r, column=3, value=a.get("nume", ""))
            cota = a.get("cota", "")
            cell_d = ws.cell(row=r, column=4, value=cota)
            if isinstance(cota, (int, float)):
                cell_d.number_format = "0.00%"
            ws.cell(row=r, column=5, value=comp.get("administrator", "") if j == 0 else "")
            ws.cell(row=r, column=6, value=comp.get("relatii", "") if j == 0 else "")
            ws.cell(row=r, column=7, value=comp.get("adresa", "") if j == 0 else "")
            _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
            r += 1
        # Merge company name / admin / relatii / adresa pe randurile asociatilor
        if len(asociati) > 1:
            for col in (2, 5, 6, 7):
                ws.merge_cells(start_row=first, start_column=col, end_row=r - 1, end_column=col)
        # CAEN principal
        ws.cell(row=r, column=2, value="CAEN principal")
        ws.cell(row=r, column=3, value=comp.get("caen_principal", ""))
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
        r += 1
        # CAEN preponderent
        ws.cell(row=r, column=2, value="CAEN preponderent")
        ws.cell(row=r, column=3, value=comp.get("caen_preponderent", ""))
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
        r += 2

    # -------------------- 2. Verdict --------------------
    ws.cell(row=r, column=2, value="2. CONCLUZIE LEGATURI").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    ws.cell(row=r, column=2, value=verdict.get("tip", ""))
    ws.cell(row=r, column=3, value=verdict.get("cu", ""))
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    _style_row(ws, r, 2, 7, fill=_VERDICT_FILL, font=_BOLD, align=_WRAP)
    r += 1
    if verdict.get("nota"):
        ws.cell(row=r, column=2, value="Nota")
        ws.cell(row=r, column=3, value=verdict["nota"])
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        _style_row(ws, r, 2, 7, font=Font(italic=True, size=9), align=_WRAP)
        r += 1
    r += 1

    # -------------------- 3. Tabele financiare per an --------------------
    ws.cell(row=r, column=2, value="3. DATE FINANCIARE CONSOLIDATE (per an)").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    FIN_HDRS = ["Denumire societate", "Nr. angajati", "CA (lei)",
                "Active totale (lei)", "CA (euro)", "Active totale (euro)"]
    for yd in years:
        an = yd["an"]
        curs = yd["curs"]
        # Antet an + curs
        ws.cell(row=r, column=2, value=f"Exercitiul financiar {an}").font = _BOLD
        ws.cell(row=r, column=9, value="curs euro").font = _NORMAL
        jcell = ws.cell(row=r, column=10, value=curs)
        jcell.number_format = "0.0000"
        curs_ref = f"$J${r}"
        r += 1
        # Header tabel
        for i, h in enumerate(FIN_HDRS):
            ws.cell(row=r, column=2 + i, value=h)
        _style_row(ws, r, 2, 7, fill=_SUBHDR_FILL, font=_BOLD, align=_CENTER)
        r += 1
        data_first = r
        for row in yd["rows"]:
            ws.cell(row=r, column=2, value=row["denumire"] + (" [EXCLUS]" if row.get("exclus") else ""))
            ang = row.get("angajati", "-")
            ws.cell(row=r, column=3, value=ang)
            cl = ws.cell(row=r, column=4, value=row.get("ca_lei", 0))
            cl.number_format = "#,##0.00"
            al = ws.cell(row=r, column=5, value=row.get("active_lei", 0))
            al.number_format = "#,##0.00"
            fe = ws.cell(row=r, column=6, value=f"=D{r}/{curs_ref}")
            fe.number_format = "#,##0.00"
            ge = ws.cell(row=r, column=7, value=f"=E{r}/{curs_ref}")
            ge.number_format = "#,##0.00"
            _style_row(ws, r, 2, 7, font=_NORMAL, align=_WRAP)
            if row.get("nota"):
                ws.cell(row=r, column=9, value=row["nota"]).font = Font(italic=True, size=8)
            r += 1
        data_last = r - 1
        # TOTAL — sumeaza doar randurile neexcluse
        incl_rows = [data_first + i for i, row in enumerate(yd["rows"]) if not row.get("exclus")]
        ws.cell(row=r, column=2, value="TOTAL (intreprinderi luate in calcul)")
        if incl_rows:
            c_terms = "+".join(f"C{rr}" for rr in incl_rows)
            # angajati pot fi "-"; folosim SUM pe range si ignoram textul
            ws.cell(row=r, column=3, value=f"=SUM(C{data_first}:C{data_last})")
            for col_letter in ("D", "E", "F", "G"):
                terms = "+".join(f"{col_letter}{rr}" for rr in incl_rows)
                cc = ws.cell(row=r, column={"D": 4, "E": 5, "F": 6, "G": 7}[col_letter],
                             value=f"={terms}")
                cc.number_format = "#,##0.00"
        _style_row(ws, r, 2, 7, fill=_TOTAL_FILL, font=_BOLD, align=_CENTER)
        r += 2

    # -------------------- 4. Categoria finala --------------------
    ws.cell(row=r, column=2, value="4. CATEGORIA IMM").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    ws.cell(row=r, column=2, value="Incadrare finala:")
    cell_cat = ws.cell(row=r, column=3, value=categorie)
    cell_cat.font = Font(bold=True, size=12, color="C00000")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    _style_row(ws, r, 2, 5, font=_BOLD, align=_CENTER)

    ws.sheet_view.showGridLines = False
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    # Smoke minimal
    out = Path.cwd() / "_smoke_analiza_imm.xlsx"
    build_analiza_imm(
        out,
        {"denumire": "DEMO SRL", "cui": "12345678"},
        [{"denumire": "DEMO SRL", "cui": "12345678",
          "asociati": [{"nume": "POP ION", "cota": 1.0}],
          "administrator": "POP ION", "relatii": "-",
          "adresa": "Bucuresti", "caen_principal": "6201", "caen_preponderent": "6201"}],
        {"tip": "AUTONOMA", "cu": "-", "nota": "demo"},
        [{"an": 2024, "curs": 4.9741,
          "rows": [{"denumire": "DEMO SRL", "angajati": 5, "ca_lei": 1000000, "active_lei": 500000}]}],
        "MICROINTREPRINDERE",
    )
    print(f"OK: {out}")
