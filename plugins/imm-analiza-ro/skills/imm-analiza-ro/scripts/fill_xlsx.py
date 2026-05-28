"""Completare Declaratie IMM.xlsx — DOAR Date_partenere și Date_legate.

INVARIANTE:
- Niciodată salva peste template-ul original
- Niciodată atinge row 12 din Date_partenere (formule)
- Niciodată atinge sheet-urile: Ipoteze, Date_consolidate, Calcul partenere & legate,
  Sect A - Intr. partenere, Tabel B1 - consolidate, Tabel B2 - legate
- Procent în coloana G = decimal (0.30), format 0.00% îl afișează ca 30%
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from _paths import analysis_dir, template_path

MAX_PARTENERI = 8
MAX_LEGATE = 10  # template suportă rândurile 4-13 în Date_legate


def fill_excel(
    parteneri: list[dict],
    legate: list[dict],
    solicitant_cui: str,
    cwd: Path | None = None,
    date_override: date | None = None,
) -> Path:
    """Completează Excel-ul și returnează path-ul fișierului nou creat.

    `parteneri` items au cheile: denumire, adresa, cui, reprezentant, procent,
                                 salariati, cifra_afaceri_mii_lei, active_totale_mii_lei
    `legate` items au cheile: denumire, salariati, cifra_afaceri_mii_lei, active_totale_mii_lei
    """
    if len(parteneri) > MAX_PARTENERI:
        raise ValueError(
            f"Template-ul suportă max {MAX_PARTENERI} parteneri; primit {len(parteneri)}"
        )
    if len(legate) > MAX_LEGATE:
        raise ValueError(
            f"Template-ul suportă max {MAX_LEGATE} legate; primit {len(legate)}"
        )

    template = template_path("Declaratie_IMM_v1.0.xlsx")
    out_dir = analysis_dir(solicitant_cui, cwd, date_override)
    out_path = out_dir / f"03_Declaratie_IMM_{solicitant_cui}_completata.xlsx"

    shutil.copy(template, out_path)
    wb = load_workbook(out_path)

    # ---------- Sheet 2: Date_partenere ----------
    ws = wb["Date_partenere"]
    for i, p in enumerate(parteneri):
        r = 4 + i
        ws.cell(row=r, column=3, value=p.get("denumire", ""))
        ws.cell(row=r, column=4, value=p.get("adresa", ""))
        ws.cell(row=r, column=5, value=str(p.get("cui", "")))
        ws.cell(row=r, column=6, value=p.get("reprezentant", ""))

        cell_g = ws.cell(row=r, column=7, value=float(p.get("procent", 0)) / 100.0)
        cell_g.number_format = "0.00%"

        cell_h = ws.cell(row=r, column=8, value=float(p.get("salariati", 0)))
        cell_h.number_format = "#,##0.00"

        cell_i = ws.cell(row=r, column=9, value=float(p.get("cifra_afaceri_mii_lei", 0)))
        cell_i.number_format = "#,##0.000"

        cell_j = ws.cell(row=r, column=10, value=float(p.get("active_totale_mii_lei", 0)))
        cell_j.number_format = "#,##0.000"

    # ---------- Sheet 4: Date_legate (dacă există legate) ----------
    if legate and "Date_legate" in wb.sheetnames:
        wl = wb["Date_legate"]
        for i, l in enumerate(legate):
            r = 4 + i
            wl.cell(row=r, column=3, value=l.get("denumire", ""))
            cell_d = wl.cell(row=r, column=4, value=float(l.get("salariati", 0)))
            cell_d.number_format = "#,##0.00"
            cell_e = wl.cell(row=r, column=5, value=float(l.get("cifra_afaceri_mii_lei", 0)))
            cell_e.number_format = "#,##0.000"
            cell_f = wl.cell(row=r, column=6, value=float(l.get("active_totale_mii_lei", 0)))
            cell_f.number_format = "#,##0.000"

    wb.save(out_path)
    return out_path


# -------------------- Smoke test --------------------

def _smoke():
    parteneri = [
        {
            "denumire": "ALPHA SRL", "adresa": "Str. Test 1, București",
            "cui": "12345678", "reprezentant": "POPESCU ION",
            "procent": 30.0, "salariati": 25,
            "cifra_afaceri_mii_lei": 4500.123, "active_totale_mii_lei": 3200.456,
        },
        {
            "denumire": "BETA SA", "adresa": "Str. Demo 5, Cluj",
            "cui": "23456789", "reprezentant": "IONESCU MARIA",
            "procent": 40.0, "salariati": 80,
            "cifra_afaceri_mii_lei": 12300.789, "active_totale_mii_lei": 9876.543,
        },
    ]
    legate = [
        {
            "denumire": "GAMMA SRL (deținută 60%)",
            "salariati": 15,
            "cifra_afaceri_mii_lei": 2100.000,
            "active_totale_mii_lei": 1500.000,
        }
    ]
    out = fill_excel(parteneri, legate, "TESTCUI_99999",
                     cwd=Path(__file__).parent.parent / "_test_output")
    print(f"OK — Excel completat la: {out}")
    print(f"     Size: {out.stat().st_size} bytes")


if __name__ == "__main__":
    _smoke()
